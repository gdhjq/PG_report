from openpyxl import load_workbook,worksheet
from openpyxl.styles import Alignment,Font,Border, Side, colors
import datetime,time,pytz
import os
import random
from docxtpl import DocxTemplate
import http.client

def get_webservertime():
    time_conn=http.client.HTTPConnection('www.baidu.com')
    time_conn.request("GET", "/")
    r=time_conn.getresponse()
    #r.getheaders() #获取所有的http头
    ts=  r.getheader('date') #获取http头date部分
    #将GMT时间转换成北京时间
    ltime= time.strptime(ts[5:25], "%d %b %Y %H:%M:%S")
    # print(ltime)
    ttime=time.localtime(time.mktime(ltime)+8*60*60)
    # print(ttime)
    dat="%u%02u%02u"%(ttime.tm_year,ttime.tm_mon,ttime.tm_mday)
    tm="%02u:%02u:%02u"%(ttime.tm_hour,ttime.tm_min,ttime.tm_sec)
    currenttime=dat

    date_list = time.strptime(dat, "%Y%m%d")
    y, m, d = date_list[:3]
    delta = datetime.timedelta(days=1)
    date_result = datetime.datetime(y, m, d) + delta
    tomorrow_day = date_result.strftime("%Y-%m-%d")

    y,m,d = time.strptime(tomorrow_day,"%Y-%m-%d")[:3]
    write_date_format2 = str(y)+'年' +str(m)+'月'+str(d)+'日'
    return currenttime,tomorrow_day,write_date_format2,y,m,d

#目录
dir_path = os.path.dirname(os.path.realpath(__file__))
template_file_name = 'fruit.docx'

##字体
fontObj1 = Font(name=u'微软雅黑', bold=True, italic=True, size=12)
fontObj2 = Font(name=u'楷体', bold=True, size=12)
fontObj3 = Font(name=u'宋体', size=12)
border_set = Border(left=Side(style='thin', color=colors.BLACK),
                    right=Side(style='thin', color=colors.BLACK),
                    top=Side(style='thin', color=colors.BLACK),
                    bottom=Side(style='thin', color=colors.BLACK))

#时间
write_date,tomorow_date,write_date_format2,y,m,d = get_webservertime()
print("今天",write_date)
print("明天",tomorow_date)
# write_date = time.strftime("%Y%m%d", time.localtime())
# print(write_date)
# tomorow_date = (datetime.datetime.now()+datetime.timedelta(days=1)).strftime('%Y-%m-%d')
# year_format =  (datetime.datetime.now()+datetime.timedelta(days=1)).year
# month_format =  (datetime.datetime.now()+datetime.timedelta(days=1)).month
# day_format =  (datetime.datetime.now()+datetime.timedelta(days=1)).day
# write_date_format2 = str(year_format)+'年' +str(month_format)+'月'+str(day_format)+'日'

#others
choose_sheet = "送货"
source_sheet = "Data"
factory2id={"2000":"长隆集团有限公司","2100":"香江野生动物世界分公司","2101":"鳄鱼公园分公司","2102":"长隆夜间动物世界分公司欢乐世界","2103":"长隆夜间动物世界分公司大马戏","2104":"XXXXX","2105":"香江酒家分公司","2106":"香江酒店分公司","2107":"广州长隆酒店分公司","2108":"熊猫酒店分公司","2400":"香江房地产发展有限公司"}

#execl格式设置
page_split_num = 45
page_num=50

class ExcelOp(object):
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.active
        # self.sheet = sheets[0]
        # self.ws = self.wb[self.sheet]
        self.ws = self.wb[source_sheet]

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 是否存在已有sheet，存在就删除重写
    def check_sheet(self,choose_sheet):
        print("检查choosesheet")
        if choose_sheet in self.wb.sheetnames:
            ws_exist = self.wb[choose_sheet]
            self.wb.remove(ws_exist)
        self.wb.create_sheet(choose_sheet)

    def change_ws(self,change_sheet):
        self.ws = self.wb[change_sheet]

    # 获取指定表单sheet_name全部值
    def get_all_value(self,sheet_name=source_sheet):
        ws_choose = self.wb[sheet_name]
        back_list = []
        for row in ws_choose.iter_rows():
            list_row = []
            for cell in row:
                list_row.append(cell.value)
            back_list.append(list_row)
        return back_list


def handle_fruit_hotel(tomorow_data_ls):
    """
    根据明天要送货的数据，返回全部的水果和各酒店水果情况
    :param tomorow_data_ls:
    :return:
    """
    ##按hotel去区分
    site_dict = {}              #{'hotel':{"partment":{"白虎":[[][]]},"hotel_fruit":{"xx":{"num":[20,30],"unit":["KG"]}} },}
    fruit_dict= {}              #{'1010500066': {'name': 'xx', 'sum': ['40', '20', '80']},}
    for row_ls in tomorow_data_ls:
        ##所有酒店fruit汇总
        if fruit_dict.get(row_ls[index_of_fruit_num]):
            fruit_dict[row_ls[index_of_fruit_num]]["sum"].append(float(row_ls[index_of_want_num]))
        else:
            fruit_dict[row_ls[index_of_fruit_num]] = {"name": row_ls[index_of_fruit_desc], "sum": [float(row_ls[index_of_want_num])], "unit": row_ls[index_of_count_size],}

        ##按照酒店去划分部门和水果
        if site_dict.get(row_ls[index_of_hotel]):
            # print(row_ls[index_of_partment])
            if site_dict[row_ls[index_of_hotel]]["partment"].get(row_ls[index_of_partment]):
                site_dict[row_ls[index_of_hotel]]["partment"][row_ls[index_of_partment]].append([row_ls[index_of_fruit_desc],row_ls[index_of_want_num],row_ls[index_of_count_size]])
            else:
                site_dict[row_ls[index_of_hotel]]["partment"][row_ls[index_of_partment]] = [[row_ls[index_of_fruit_desc],row_ls[index_of_want_num],row_ls[index_of_count_size]]]

            if site_dict[row_ls[index_of_hotel]]["hotel_fruit"].get(row_ls[index_of_fruit_num]):
                site_dict[row_ls[index_of_hotel]]["hotel_fruit"][row_ls[index_of_fruit_num]]["num"].append(row_ls[index_of_want_num])
            else:
                site_dict[row_ls[index_of_hotel]]["hotel_fruit"][row_ls[index_of_fruit_num]]={
                                                        "num":[row_ls[index_of_want_num]],
                                                        "unit":row_ls[index_of_count_size],
                                                        "name":row_ls[index_of_fruit_desc]}
        else:
            site_dict[row_ls[index_of_hotel]] = {"partment":{
                                                    row_ls[index_of_partment]:
                                                        [[row_ls[index_of_fruit_desc],row_ls[index_of_want_num],row_ls[index_of_count_size]]]} ,
                                                 "hotel_fruit":{
                                                    row_ls[index_of_fruit_num]:{
                                                        "num":[row_ls[index_of_want_num]],
                                                        "unit":row_ls[index_of_count_size],
                                                        "name":row_ls[index_of_fruit_desc]}
                                                    }
                                                 }

    return fruit_dict,site_dict


def check_date_dir(dir_name):
    """"""
    trans_dir = os.path.join(dir_path, dir_name)
    trans_dir2 = os.path.join(dir_path, dir_name, write_date)
    if not os.path.exists(trans_dir2):
        if not os.path.exists(trans_dir):
            os.mkdir(trans_dir)
        os.mkdir(trans_dir2)

def init_env(dir_name):
    '''初始化找文件和路径'''
    ##根据xlsx类型找文件

    read_dir = os.path.join(dir_path,dir_name, write_date)
    file_dir = []
    print(read_dir)
    if os.path.exists(read_dir):
        for root, dirs, files in os.walk(read_dir):
            # print("共有单数",len(files))
            for file in files:
                filename,type = os.path.splitext(file)
                if type == '.xlsx':
                    file_excel = os.path.join(dir_path,dir_name, write_date,filename) + ".xlsx"
                    print(file_excel)
                    file_dir.append(file_excel)
    else:
        print('没有日期目录')
    return file_dir


if __name__ == "__main__":
    check_date_dir("JYZ")
    check_date_dir("B")
    file_name_ls = init_env("A")
    print(file_name_ls)
    excel_file = file_name_ls[0]
    print(excel_file)
    excel_object = ExcelOp(file=excel_file)
    r,c = excel_object.get_row_clo_num()
    print("列和行",c,r)

    # 根据列的数字返回字母
    # print(get_column_letter(index_of_hotel+1))

    #是否存在已有sheet，存在就删除重写
    excel_object.check_sheet(choose_sheet)

    #去获得source_sheet所有值
    back_ls = excel_object.get_all_value(source_sheet)


    #获取标题头
    header_ls = back_ls[0]
    index_of_bill_num = header_ls.index("采购凭证")
    index_of_hotel = header_ls.index("工厂")
    index_of_partment = header_ls.index("需求部门名称")
    index_of_fruit_num = header_ls.index("物料")
    index_of_fruit_desc = header_ls.index("物料描述")
    index_of_diver_date = header_ls.index("交货日期")
    index_of_divered_num = header_ls.index("已交货数量")
    index_of_count_size = header_ls.index("基本计量单位")
    index_of_want_num = header_ls.index("数量-基本计量单位")
    index_of_beizhu_num = header_ls.index("数量-基本计量单位")
    ##把今天和非今天的分开
    tomorow_data_ls = []
    not_tomorow_data = []
    for data_ls in back_ls[1:]:
        diver_date = data_ls[index_of_diver_date].date().strftime('%Y-%m-%d')
        if tomorow_date != diver_date:
            not_tomorow_data.append(data_ls)
        else:

            tomorow_data_ls.append(data_ls)

    ####开始计算
    fruit_dict,site_dict = handle_fruit_hotel(tomorow_data_ls)


    #切换sheet
    excel_object.change_ws(choose_sheet)
    ws = excel_object.ws

    #test查看
    # for hotel,values in site_dict.items():
    #     print("*"*10,factory2id[hotel])
    #     for part_fruit_type,vls in values.items():
    #         if part_fruit_type=="hotel_fruit":
    #             print("-"*10,"共需要：")
    #         for part,fruits in vls.items():
    #             if part_fruit_type == "partment":
    #                 print("部门：",part)
    #                 for v in fruits:
    #                     print(v)
    #             else:
    #                 print(fruits["name"],sum(fruits["num"]),fruits["unit"])

    ##计算长度
    hotel_row_long_num = {}
    for hotel_num, values in site_dict.items():
        hotel_row_long_num[hotel_num]=1
        for part_fruit_type, vls in values.items():
            for part, fruits in vls.items():
                if part_fruit_type == "partment":
                    hotel_row_long_num[hotel_num] +=1
                    for v in fruits:
                        hotel_row_long_num[hotel_num] +=1
    # print(hotel_row_long_num)
    #排序一下
    sort_hotel = dict(sorted(hotel_row_long_num.items(), key=lambda e: e[1]))
    print(sort_hotel)

    #将排序的再进行按页面设置分开
    excel_hotel_ls = []
    v = 0
    for h,val in sort_hotel.items():
        v = v+val
        if v >= page_split_num:
            excel_hotel_ls.append([h])
            v = val
            if val > page_num:
                excel_hotel_ls.append([])
        else:
            if excel_hotel_ls:
                excel_hotel_ls[-1].append(h)
            else:
                excel_hotel_ls.append([h])
    # print(excel_hotel_ls)


    JYZ_page_ls = []   ##检疫证列表
    ##写excel单元格
    write_row_num = 1
    write_row_times = 1
    print(site_dict)
    for i,hotel_ls in enumerate(excel_hotel_ls):
        range_begin = (i) * 50 + 1
        hotel_head_site = (i) * 50 + 1
        # print(i,hotel_ls)
        ws.cell(row=range_begin, column=2, value=tomorow_date).font = fontObj1
        for hotel_num in hotel_ls:
            ws.cell(row=range_begin, column=1, value=factory2id[hotel_num]).font = fontObj1
            # ws.cell(row=range_begin, column=1)
            range_begin += 1
            hotel_head_site = range_begin+1
            ##写部门和水果
            for site,fruits_ls in site_dict[hotel_num]["partment"].items():
                site_value = "---"+site+"---"
                ws.cell(row=range_begin,column=1,value=site_value).font = fontObj2
                range_begin+=1
                for f in fruits_ls:
                    ws.cell(row=range_begin, column=1, value=f[0]).font = fontObj3
                    ws.cell(row=range_begin, column=3, value=f[1]).font = fontObj3
                    ws.cell(row=range_begin, column=4, value=f[2]).font = fontObj3
                    range_begin+=1

            ##写各酒店的水果总计
            ws.cell(row=hotel_head_site, column=6, value=factory2id[hotel_num]).font = fontObj2
            ws.cell(row=hotel_head_site, column=6).alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row=hotel_head_site, start_column=6, end_row=hotel_head_site, end_column=9)
            ws.cell(row=hotel_head_site, column=6).border = border_set
            ws.cell(row=hotel_head_site, column=7).border = border_set
            ws.cell(row=hotel_head_site, column=8).border = border_set
            ws.cell(row=hotel_head_site, column=9).border = border_set

            render_dict = {
                'all_paragraph': [],
                'page_name':'',
                'tomo_day':[y,m,d]
            }

            hotel_head_site+=1
            for fruit_num, fruits_desc in site_dict[hotel_num]["hotel_fruit"].items():
                #先给page写上酒店名字
                # render_dict["page_name"]=factory2id[hotel_num]

                sum_fruit = sum(fruits_desc["num"])
                fruit_name = fruits_desc["name"]
                fruits_unit = fruits_desc["unit"]
                ws.cell(row=hotel_head_site, column=6, value=fruit_name).font = fontObj3
                ws.cell(row=hotel_head_site, column=6).border = border_set
                ws.cell(row=hotel_head_site, column=7).border = border_set
                ws.cell(row=hotel_head_site, column=8, value=sum_fruit).font = fontObj3
                ws.cell(row=hotel_head_site, column=8).border = border_set
                ws.cell(row=hotel_head_site, column=9, value=fruits_unit).font = fontObj3
                ws.cell(row=hotel_head_site, column=9).border = border_set
                hotel_head_site +=1

                if len(render_dict["all_paragraph"]) >17:
                    JYZ_page_ls.append(render_dict)
                    ra = str(factory2id[hotel_num]) + str(random.randint(1,1000))
                    print(ra)
                    render_dict = {'all_paragraph': [],"page_name":ra,'tomo_day':[y,m,d]}
                random_num = random.uniform(8, 18)
                fruit_name_2 = fruits_desc["name"].split(" ")[0]
                vv = [fruit_name_2, str(write_date_format2), str(round(random_num, 2)) + '%', "1公斤","合格"]
                render_dict['all_paragraph'].append(vv)
                if not render_dict["page_name"]:
                    render_dict["page_name"] = factory2id[hotel_num]
            range_begin+=1
            JYZ_page_ls.append(render_dict)

    # print(JYZ_page_ls,len(JYZ_page_ls))
    for JYZ_PAGE in JYZ_page_ls:
        print(JYZ_PAGE["page_name"])
        template_file = os.path.join(dir_path,'W', template_file_name)
        tpl = DocxTemplate(template_file)
        tpl.render(JYZ_PAGE)
        write_word_file = os.path.join('JYZ', write_date, JYZ_PAGE["page_name"]) + '检疫报告.docx'
        tpl.save(write_word_file)




    ##修改excel的表格行高
    ws.column_dimensions['A'].width = 25.0
    ws.column_dimensions['B'].width = 8.0
    ws.column_dimensions['C'].width = 6.0
    ws.column_dimensions['D'].width = 4.0
    ws.column_dimensions['E'].width = 10.0
    ws.column_dimensions['F'].width = 25.0
    ws.column_dimensions['G'].width = 4.0
    ws.column_dimensions['H'].width = 6.0
    ws.column_dimensions['I'].width = 4.0
    for i in range(1, 500):
        ws.row_dimensions[i].height = 14.65
    ws.page_margins = worksheet.page.PageMargins(top=0.48, header=0.97,right=0.37,left=0.33)

    txtfile = os.path.join(dir_path,"B",write_date,write_date)+"总计水果.txt"
    with open(txtfile,"w") as wt:
        for fruits_num,f_vals in fruit_dict.items():
            sum_num = sum(f.vals["sum"])
            w_str = f.vals["name"]+" " + str(sum_num) + "\n"
            wt.write(w_str)


    write_excel_file=os.path.join('B', write_date,write_date)+"送货.xlsx"
    excel_object.wb.save(write_excel_file)
    time.sleep(2)
